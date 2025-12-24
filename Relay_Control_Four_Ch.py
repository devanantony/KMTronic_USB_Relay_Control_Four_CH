import ttkbootstrap as tb
from ttkbootstrap.constants import *
from tkinter import messagebox
from tkinter import ttk
import serial
import serial.tools.list_ports
import openpyxl
import os

# ===========================
# USB Relay Functions
# ===========================

def list_com_ports():
    """Return a list of available COM ports."""
    ports = serial.tools.list_ports.comports()
    return [port.device for port in ports]

def send_data(data_bytes):
    port = selected_port.get()
    if not port:
        messagebox.showerror("Error", "No COM port selected.")
        return
    try:
        with serial.Serial(port, baudrate=9600, timeout=1) as ser:
            ser.write(data_bytes)
    except serial.SerialException as e:
        messagebox.showerror("Serial Error", str(e))
        reset_all_relays_log()
        for i in range(4):
            update_circle(i, "red")
            relay_buttons[i].config(text=f"Turn ON RL{i + 1}")

def reset_all_relays_log():
    file_name = "relay_log.xlsx"
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relay Log"
        wb.save(file_name)

    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    for row in range(1, 5):
        ws[f"A{row}"] = 0
    wb.save(file_name)

def write_to_log(value, pos):
    file_name = "relay_log.xlsx"
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relay Log"
        wb.save(file_name)
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    ws[pos] = value
    wb.save(file_name)

def read_from_log(pos):
    file_name = "relay_log.xlsx"
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        return ws[pos].value
    return 0

def read_relay_states():
    file_name = "relay_log.xlsx"
    values = []
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        for row in range(1, 5):
            values.append(ws[f"A{row}"].value or 0)
    else:
        values = [0, 0, 0, 0]
    return values

def save_description(i):
    file_name = "relay_log.xlsx"
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(file_name)

    if "Descriptions" not in wb.sheetnames:
        ws = wb.create_sheet("Descriptions")
    else:
        ws = wb["Descriptions"]

    if ws["A1"].value != "Relay":
        ws["A1"] = "Relay"
    if ws["B1"].value != "Description":
        ws["B1"] = "Description"

    ws[f"A{i+2}"] = f"RL{i+1}"
    ws[f"B{i+2}"] = description_entries[i].get()
    wb.save(file_name)
    messagebox.showinfo("Success", f"Description for RL{i+1} saved successfully.")

def load_descriptions():
    file_name = "relay_log.xlsx"
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
        if "Descriptions" in wb.sheetnames:
            ws = wb["Descriptions"]
            for i in range(4):
                desc = ws[f"B{i+2}"].value
                if desc:
                    description_entries[i].delete(0, 'end')
                    description_entries[i].insert(0, desc)

def clear_description(i):
    description_entries[i].delete(0, 'end')

def reset_all_relays_and_ui():
    for i in range(4):
        send_data(bytes.fromhex(f'FF 0{i + 1} 00'))
        update_circle(i, "red")
        write_to_log(0, f'A{i + 1}')
        relay_buttons[i].config(text=f"Turn ON RL{i + 1}")

def set_all_relays_and_ui():
    for i in range(4):
        send_data(bytes.fromhex(f'FF 0{i + 1} 01'))
        update_circle(i, "green")
        write_to_log(1, f'A{i + 1}')
        relay_buttons[i].config(text=f"Turn OFF RL{i + 1}")

def update_circle(index, color):
    canvas.itemconfig(circles[index], fill=color)

def toggle_relay(index):
    current_state = read_from_log(f'A{index + 1}')
    if current_state == 1:
        send_data(bytes.fromhex(f'FF 0{index + 1} 00'))
        update_circle(index, "red")
        write_to_log(0, f'A{index + 1}')
        relay_buttons[index].config(text=f"Turn ON RL{index + 1}")
    else:
        send_data(bytes.fromhex(f'FF 0{index + 1} 01'))
        update_circle(index, "green")
        write_to_log(1, f'A{index + 1}')
        relay_buttons[index].config(text=f"Turn OFF RL{index + 1}")

def refresh_com_ports():
    """Refresh the COM port list in the dropdown."""
    ports = list_com_ports()
    com_port_cb['values'] = ports
    if ports:
        selected_port.set(ports[0])
    else:
        selected_port.set('')

# ===========================
# UI Setup (Modern Theme)
# ===========================

root = tb.Window(themename="flatly")
root.title("KMTronic 4-Channel Relay Control")

# COM port selection
selected_port = tb.StringVar()
ttk.Label(root, text="Select COM Port:").grid(row=0, column=0, padx=10, pady=10, sticky='w')
com_port_cb = ttk.Combobox(root, textvariable=selected_port, values=list_com_ports(), state='readonly', width=15)
com_port_cb.grid(row=0, column=1, padx=5, pady=10)
ttk.Button(root, text="Refresh", command=refresh_com_ports).grid(row=0, column=2, padx=5)

# Relay buttons and descriptions
relay_buttons = []
description_entries = []

for i in range(4):
    btn = tb.Button(root, text="", bootstyle="success-outline", command=lambda idx=i: toggle_relay(idx))
    btn.grid(row=i+1, column=1, pady=10)
    relay_buttons.append(btn)

    tb.Label(root, text=f"RL{i+1} Description:").grid(row=i+1, column=3, sticky='e', padx=5)
    entry = tb.Entry(root, width=30)
    entry.grid(row=i+1, column=4, padx=5)
    description_entries.append(entry)

    tb.Button(root, text="Save", bootstyle="primary-outline", width=6, command=lambda idx=i: save_description(idx)).grid(row=i+1, column=5, padx=5)
    tb.Button(root, text="Clear", bootstyle="danger-outline", width=6, command=lambda idx=i: clear_description(idx)).grid(row=i+1, column=6, padx=5)

# Relay status indicator
canvas = tb.Canvas(root, width=380, height=80)
canvas.grid(row=5, column=0, columnspan=7, pady=20)
circles = []
for i in range(4):
    x = 30 + i * 85
    circle = canvas.create_oval(x, 20, x + 30, 50, fill="red")
    canvas.create_text(x + 15, 60, text=f"RL{i+1}", font=("Arial", 10))
    circles.append(circle)

# Control buttons at bottom center
control_frame = tb.Frame(root)
control_frame.grid(row=6, column=0, columnspan=7, pady=15)
tb.Button(control_frame, text="Set All", bootstyle="success", command=set_all_relays_and_ui).pack(side=LEFT, padx=10)
tb.Button(control_frame, text="Reset All", bootstyle="danger", command=reset_all_relays_and_ui).pack(side=LEFT, padx=10)

# Initialize relay states
relay_states = read_relay_states()
for i, state in enumerate(relay_states):
    if state == 1:
        update_circle(i, "green")
        relay_buttons[i].config(text=f"Turn OFF RL{i + 1}")
    else:
        update_circle(i, "red")
        relay_buttons[i].config(text=f"Turn ON RL{i + 1}")

load_descriptions()

root.mainloop()
